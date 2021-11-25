from re import sub
import openpyxl
import regex
from openpyxl.styles import Font
import paths

path = paths.excelPath
wb_obj = openpyxl.load_workbook(path) 
wb = openpyxl.Workbook() 
sheet = wb.active 

row = 1
column = 1
font = Font(size=15) # font size

# insert excel funct
def insert_to_excel(data, my_row, my_column):
    temp = sheet.cell(my_row, my_column)
    temp.value = data

# Column Numbers
company_Name_Column = 1
client_Name_Column = 2
mail_Column = 3
domain_Column = 4
zip_Code_Column = 5
phone_Number_Column = 6
invoice_Total_Column = 7
sub_Total_Column = 8
tax_Column = 9
tax_Rate_Column = 10
discount_Value_Column = 11
accountHolder_Column = 12

# title cell possitions
companyName = sheet.cell(row = 1, column = 1) # company name column title
clientName = sheet.cell(row = 1, column = 2) # client name column title
mail = sheet.cell(row = 1, column = 3) # mail column title
domain = sheet.cell(row = 1, column = 4) # domain column title
zipCode = sheet.cell(row = 1, column = 5) # zip code column title
phoneNumber = sheet.cell(row = 1, column = 6) # phone number column title
invoiceTotal = sheet.cell(row = 1, column = 7) # invoice total column title
subTotal = sheet.cell(row = 1, column = 8) # sub total column title
tax = sheet.cell(row = 1, column = 9) # tax column title
taxRate = sheet.cell(row = 1, column = 10) # tax rate column title
discount = sheet.cell(row = 1, column = 11) # discount column title
accountHolder = sheet.cell(row = 1, column = 12) # account holder column title

# title names
companyName.value = "Company Name"
clientName.value = "Client Name"
mail.value = "Mail"
domain.value = "Domain"
zipCode.value = "Zip Code"
phoneNumber.value = "Phone Number"
invoiceTotal.value = "Invoice Total"
subTotal.value = "Sub Total"
tax.value = "Tax"
taxRate.value = "Tax Rate"
discount.value = "Discount Value"
accountHolder.value = "Account Holder"

# sets title fonts
companyName.font = font
clientName.font = font
mail.font = font
domain.font = font
zipCode.font = font
phoneNumber.font = font
invoiceTotal.font = font
subTotal.font = font
tax.font = font
taxRate.font = font
discount.font = font
accountHolder.font = font

def start():
    k = 0
    new_row = 2
    # inserting values to excel
    while(new_row < 12):
        insert_to_excel(regex.find_allCompanyName()[k],new_row,company_Name_Column)
        insert_to_excel(regex.find_allClientName()[k],new_row,client_Name_Column)
        insert_to_excel(regex.find_allMail()[k],new_row,mail_Column)
        insert_to_excel(regex.find_allDomain()[k],new_row,domain_Column)
        insert_to_excel(regex.find_allZipCode()[k],new_row,zip_Code_Column)
        insert_to_excel(regex.find_allPhoneNumbers()[k],new_row,phone_Number_Column)
        insert_to_excel(regex.find_allInvoiceTotal()[k],new_row,invoice_Total_Column)
        insert_to_excel(regex.find_allSubTotal()[k],new_row,sub_Total_Column)
        insert_to_excel(regex.find_allTax()[k],new_row,tax_Column)
        insert_to_excel(regex.find_allTaxRates()[k],new_row,tax_Rate_Column)
        insert_to_excel(regex.find_allClientName()[k],new_row,client_Name_Column)
        insert_to_excel(regex.find_allDiscount()[k],new_row,discount_Value_Column)
        insert_to_excel(regex.find_allAccountHolders()[k],new_row,accountHolder_Column)
        new_row += 1
        k += 1
start()

# resize columns for more readable
for col in sheet.columns:
     max_length = 0
     column = col[0].column_letter # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2
     sheet.column_dimensions[column].width = adjusted_width

wb.save(path) 